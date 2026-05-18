from dotenv import load_dotenv
load_dotenv()

import json
import io
import os
import sys
import csv
import queue
import zipfile
import subprocess
import threading

from flask import (
    Flask,
    render_template,
    jsonify,
    request,
    Response,
    stream_with_context,
    send_file,
)
from werkzeug.utils import secure_filename


app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

INVOICE_FOLDER = os.path.join(BASE_DIR, "invoices")
INPUT_FOLDER = os.path.join(BASE_DIR, "Input")
OUTPUT_FOLDER = os.path.join(BASE_DIR, "Output")
CLIENTS_FOLDER = os.path.join(BASE_DIR, "Clients")

TRACKER_INPUT = os.path.join(INPUT_FOLDER, "ACCT-108 Master Invoice Tracker 2026.xlsx")
TRACKER_OUTPUT = os.path.join(OUTPUT_FOLDER, "ACCT-108 Master Invoice Tracker 2026 - Updated.xlsx")
REPORT_PATH = os.path.join(OUTPUT_FOLDER, "invoice_sort_report.csv")
CREDENTIALS = os.path.join(BASE_DIR, "credentials.json")
DRIVE_CONFIG = os.path.join(BASE_DIR, "drive_config.json")
GEMINI_STATE_PATH = os.path.join(OUTPUT_FOLDER, "gemini_state.json")

MEDIA_PLANS_FOLDER = os.path.join(BASE_DIR, "media_plans")
PRISMA_OUTPUT_DIR = os.path.join(OUTPUT_FOLDER, "prisma")
BUYING_GUIDE_PATH = os.path.join(INPUT_FOLDER, "ACCT 108 BuyingGuide.xlsx")
PRISMA_TEMPLATE = os.path.join(
    INPUT_FOLDER,
    "ACCT-108 DIGITAL TEMPLATE MEDIA PLAN IMPORT placements.xlsx",
)

ALLOWED_EXTENSIONS = {"pdf"}

os.makedirs(INVOICE_FOLDER, exist_ok=True)
os.makedirs(INPUT_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(CLIENTS_FOLDER, exist_ok=True)
os.makedirs(MEDIA_PLANS_FOLDER, exist_ok=True)
os.makedirs(PRISMA_OUTPUT_DIR, exist_ok=True)


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def allowed_plan_file(filename):
    return filename.lower().endswith((".xlsx", ".xls"))


def gemini_configured():
    if os.environ.get("GEMINI_API_KEY", "").strip():
        return True

    env_path = os.path.join(BASE_DIR, ".env")

    if os.path.exists(env_path):
        try:
            with open(env_path, encoding="utf-8") as f:
                for line in f:
                    line = line.strip()

                    if line.startswith("GEMINI_API_KEY="):
                        val = line.split("=", 1)[1].strip().strip('"').strip("'")

                        if val:
                            return True
        except Exception:
            pass

    return False


def get_gemini_enabled():
    if not os.path.exists(GEMINI_STATE_PATH):
        return True

    try:
        with open(GEMINI_STATE_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)

        return bool(data.get("enabled", True))
    except Exception:
        return True


def set_gemini_enabled(enabled):
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    with open(GEMINI_STATE_PATH, "w", encoding="utf-8") as f:
        json.dump({"enabled": bool(enabled)}, f, indent=2)


def gemini_available():
    return gemini_configured() and get_gemini_enabled()


def gemini_status_payload():
    configured = gemini_configured()
    enabled = get_gemini_enabled()

    return {
        "configured": configured,
        "enabled": enabled,
        "active": configured and enabled,
    }


def get_invoice_supplier_counts_from_report():
    counts = {
        "Meta": 0,
        "Google": 0,
        "Apple": 0,
        "AdsJoy": 0,
    }

    if not os.path.exists(REPORT_PATH):
        return counts

    try:
        seen = set()

        with open(REPORT_PATH, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)

            for row in reader:
                status = str(row.get("status", "")).strip()
                supplier = str(row.get("supplier", "")).strip()
                original_file = str(row.get("original_file", "")).strip()

                if status != "OK - copied":
                    continue

                if supplier not in counts:
                    continue

                if not original_file:
                    continue

                key = (supplier, original_file)

                if key in seen:
                    continue

                seen.add(key)
                counts[supplier] += 1

        return counts

    except Exception:
        return counts


def get_uploaded_pdf_count():
    pdf_count = 0

    if os.path.exists(INVOICE_FOLDER):
        for _, _, files in os.walk(INVOICE_FOLDER):
            pdf_count += sum(1 for f in files if f.lower().endswith(".pdf"))

    return pdf_count


def stream_script(script_name, q):
    try:
        script_path = os.path.join(BASE_DIR, script_name)
        env = os.environ.copy()
        env["GEMINI_ENABLED"] = "1" if get_gemini_enabled() else "0"

        proc = subprocess.Popen(
            [sys.executable, script_path],
            cwd=BASE_DIR,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
            bufsize=1,
            env=env,
        )

        for line in proc.stdout:
            q.put(line.rstrip())

        proc.wait()
        q.put(f"__EXIT__{proc.returncode}")

    except Exception as e:
        q.put(f"[ERR] Failed to start {script_name}: {e}")
        q.put("__EXIT__1")


def run_and_stream(script_name):
    q = queue.Queue()
    t = threading.Thread(target=stream_script, args=(script_name, q), daemon=True)
    t.start()

    def generate():
        while True:
            try:
                line = q.get(timeout=120)

                if line.startswith("__EXIT__"):
                    code = line.replace("__EXIT__", "")
                    msg = (
                        "[DONE] Script finished (exit 0)."
                        if code == "0"
                        else f"[ERR] Script exited with code {code}."
                    )
                    yield msg + "\n"
                    break

                yield line + "\n"

            except queue.Empty:
                yield "[WARN] Timeout - no output for 120s\n"
                break

    return Response(
        stream_with_context(generate()),
        mimetype="text/plain",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/gemini/status", methods=["GET"])
def api_gemini_status():
    return jsonify(gemini_status_payload())


@app.route("/api/gemini/toggle", methods=["POST"])
def api_gemini_toggle():
    data = request.get_json(silent=True) or {}
    enabled = bool(data.get("enabled", True))

    set_gemini_enabled(enabled)

    payload = gemini_status_payload()
    payload.update({
        "ok": True,
        "message": "Gemini enabled." if enabled else "Gemini disabled.",
    })

    return jsonify(payload)


@app.route("/api/status")
def status():
    supplier_counts = get_invoice_supplier_counts_from_report()
    successful_invoice_count = sum(supplier_counts.values())
    pdf_count = successful_invoice_count or get_uploaded_pdf_count()

    client_list = []

    if os.path.exists(CLIENTS_FOLDER):
        client_list = [
            d for d in os.listdir(CLIENTS_FOLDER)
            if os.path.isdir(os.path.join(CLIENTS_FOLDER, d))
        ]

    tracker_exists = os.path.exists(TRACKER_INPUT) or os.path.exists(TRACKER_OUTPUT)
    report_exists = os.path.exists(REPORT_PATH)
    creds_exists = os.path.exists(CREDENTIALS)

    folder_id_set = False

    if os.path.exists(DRIVE_CONFIG):
        try:
            with open(DRIVE_CONFIG, encoding="utf-8") as f:
                cfg = json.load(f)

            folder_id_set = bool(cfg.get("root_folder_id", "").strip())
        except Exception:
            folder_id_set = False

    gemini_state = gemini_status_payload()

    return jsonify({
        "pdf_count": pdf_count,
        "uploaded_pdf_count": get_uploaded_pdf_count(),
        "processed_invoice_count": successful_invoice_count,
        "client_folders": len(client_list),
        "client_list": client_list,
        "supplier_counts": supplier_counts,
        "tracker_exists": tracker_exists,
        "report_exists": report_exists,
        "creds_exists": creds_exists,
        "folder_id_set": folder_id_set,
        "gemini": gemini_state["active"],
        "gemini_configured": gemini_state["configured"],
        "gemini_enabled": gemini_state["enabled"],
        "gemini_active": gemini_state["active"],
    })


@app.route("/api/summary")
def summary():
    tracker_path = (
        TRACKER_OUTPUT if os.path.exists(TRACKER_OUTPUT) else
        TRACKER_INPUT if os.path.exists(TRACKER_INPUT) else
        None
    )

    if not tracker_path:
        return jsonify({
            "currency_totals": {},
            "supplier_totals": {},
            "has_data": False,
        })

    try:
        import openpyxl

        wb = openpyxl.load_workbook(tracker_path, read_only=True, data_only=True)

        currency_totals = {}
        supplier_totals = {}

        sheet_map = {
            "Adsjoy": "AdsJoy",
            "Apple (ASA)": "Apple",
            "Google": "Google",
            "Meta (facebook)": "Meta",
        }

        for sheet_name, supplier_label in sheet_map.items():
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            headers = [
                str(c.value).strip() if c.value else ""
                for c in next(ws.iter_rows(min_row=1, max_row=1))
            ]

            cur_idx = next(
                (i for i, h in enumerate(headers) if h.strip() in ("Currency", "Currency ")),
                None,
            )

            amt_idx = next(
                (i for i, h in enumerate(headers) if h.strip() == "Amount"),
                None,
            )

            if cur_idx is None or amt_idx is None:
                continue

            for row in ws.iter_rows(min_row=2, values_only=True):
                cur = str(row[cur_idx]).strip() if row[cur_idx] else ""
                amt = row[amt_idx]

                if not cur or cur == "None" or amt is None:
                    continue

                try:
                    amt = float(amt)
                except (ValueError, TypeError):
                    continue

                currency_totals[cur] = currency_totals.get(cur, 0.0) + amt
                supplier_totals[supplier_label] = supplier_totals.get(supplier_label, 0.0) + amt

        wb.close()

        return jsonify({
            "currency_totals": currency_totals,
            "supplier_totals": supplier_totals,
            "has_data": bool(currency_totals),
        })

    except Exception as e:
        return jsonify({
            "currency_totals": {},
            "supplier_totals": {},
            "has_data": False,
            "error": str(e),
        })


@app.route("/api/upload", methods=["POST"])
def upload():
    if "files" not in request.files:
        return jsonify({"error": "No files provided"}), 400

    saved = []
    errors = []

    for f in request.files.getlist("files"):
        if f and allowed_file(f.filename):
            filename = secure_filename(f.filename)
            f.save(os.path.join(INVOICE_FOLDER, filename))
            saved.append(filename)
        else:
            errors.append(f.filename)

    return jsonify({
        "message": f"Uploaded {len(saved)} file(s)" + (f", {len(errors)} skipped" if errors else ""),
        "saved": saved,
        "errors": errors,
    })


@app.route("/api/upload/tracker", methods=["POST"])
def upload_tracker():
    import shutil
    from datetime import datetime

    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    f = request.files["file"]

    if not f.filename.endswith(".xlsx"):
        return jsonify({"error": "Only .xlsx files accepted"}), 400

    if os.path.exists(TRACKER_INPUT):
        backup_name = os.path.join(
            INPUT_FOLDER,
            f"ACCT-108 Master Invoice Tracker 2026_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
        shutil.copy2(TRACKER_INPUT, backup_name)

    f.save(TRACKER_INPUT)

    return jsonify({
        "message": "[OK] Tracker uploaded to Input/ folder (previous version backed up)"
    })


@app.route("/api/upload/buying-guide", methods=["POST"])
def upload_buying_guide():
    import shutil
    from datetime import datetime

    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    f = request.files["file"]

    if not f.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Only .xlsx files accepted"}), 400

    if os.path.exists(BUYING_GUIDE_PATH):
        backup = os.path.join(
            INPUT_FOLDER,
            f"ACCT 108 BuyingGuide_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
        shutil.copy2(BUYING_GUIDE_PATH, backup)

    f.save(BUYING_GUIDE_PATH)

    try:
        from buying_guide import BuyingGuide

        bg = BuyingGuide(BUYING_GUIDE_PATH)

        return jsonify({
            "message": f"[OK] Buying Guide uploaded · {len(bg)} rows · clients: {', '.join(bg.clients())}",
            "rows": len(bg),
            "clients": bg.clients(),
            "valid": True,
        })

    except Exception as e:
        return jsonify({
            "message": f"[WARN] Saved but failed to parse: {e}",
            "error": str(e),
            "valid": False,
        }), 200


@app.route("/api/upload/prisma-template", methods=["POST"])
def upload_prisma_template():
    import shutil
    from datetime import datetime

    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    f = request.files["file"]

    if not f.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Only .xlsx files accepted"}), 400

    if os.path.exists(PRISMA_TEMPLATE):
        backup = os.path.join(
            INPUT_FOLDER,
            f"PRISMA_TEMPLATE_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
        shutil.copy2(PRISMA_TEMPLATE, backup)

    f.save(PRISMA_TEMPLATE)

    try:
        import openpyxl

        wb = openpyxl.load_workbook(PRISMA_TEMPLATE, read_only=True)
        has_sheet = "Digital import sheet ALL TYPES" in wb.sheetnames
        sheets = list(wb.sheetnames)
        wb.close()

        if not has_sheet:
            return jsonify({
                "message": f"[WARN] Saved but missing required sheet 'Digital import sheet ALL TYPES'. Found: {sheets}",
                "valid": False,
                "sheets": sheets,
            }), 200

        return jsonify({
            "message": "[OK] Prisma template uploaded and validated",
            "sheets": sheets,
            "valid": True,
        })

    except Exception as e:
        return jsonify({
            "message": f"[WARN] Saved but failed to validate: {e}",
            "error": str(e),
            "valid": False,
        }), 200


@app.route("/api/upload/credentials", methods=["POST"])
def upload_credentials():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    f = request.files["file"]

    if not f.filename.endswith(".json"):
        return jsonify({"error": "Only .json files accepted"}), 400

    f.save(CREDENTIALS)

    return jsonify({
        "message": "[OK] credentials.json saved to project root"
    })


@app.route("/api/credentials/delete", methods=["POST"])
def delete_credentials():
    if os.path.exists(CREDENTIALS):
        os.remove(CREDENTIALS)
        return jsonify({"message": "[OK] credentials.json removed"})

    return jsonify({"message": "[WARN] No credentials file found"}), 404


@app.route("/api/drive/config", methods=["GET"])
def get_drive_config():
    if not os.path.exists(DRIVE_CONFIG):
        return jsonify({"root_folder_id": ""})

    with open(DRIVE_CONFIG, encoding="utf-8") as f:
        return jsonify(json.load(f))


@app.route("/api/drive/config", methods=["POST"])
def save_drive_config():
    data = request.get_json()
    folder_id = (data or {}).get("root_folder_id", "").strip()

    if not folder_id:
        return jsonify({"error": "Folder ID cannot be empty"}), 400

    with open(DRIVE_CONFIG, "w", encoding="utf-8") as f:
        json.dump({"root_folder_id": folder_id}, f)

    return jsonify({
        "message": f"[OK] Drive folder ID saved: {folder_id}"
    })


@app.route("/api/drive/config/delete", methods=["POST"])
def delete_drive_config():
    if os.path.exists(DRIVE_CONFIG):
        os.remove(DRIVE_CONFIG)

    return jsonify({
        "message": "[OK] Drive folder ID cleared"
    })


@app.route("/api/download/report")
def download_report():
    if not os.path.exists(REPORT_PATH):
        return jsonify({"error": "Report not found"}), 404

    return send_file(
        REPORT_PATH,
        as_attachment=True,
        download_name="invoice_sort_report.csv",
    )


@app.route("/api/download/clients-zip")
def download_clients_zip():
    if not os.path.exists(CLIENTS_FOLDER):
        return jsonify({
            "error": "Clients/ folder not found. Run the workflow first."
        }), 404

    has_files = any(files for _, _, files in os.walk(CLIENTS_FOLDER))

    if not has_files:
        return jsonify({
            "error": "Clients/ folder is empty. Run the workflow first."
        }), 404

    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for dirpath, _, filenames in os.walk(CLIENTS_FOLDER):
            for filename in filenames:
                file_path = os.path.join(dirpath, filename)
                arcname = os.path.relpath(file_path, os.path.dirname(CLIENTS_FOLDER))
                zf.write(file_path, arcname)

    zip_buffer.seek(0)

    return send_file(
        zip_buffer,
        mimetype="application/zip",
        as_attachment=True,
        download_name="Clients.zip",
    )


@app.route("/api/download/tracker")
def download_tracker():
    if os.path.exists(TRACKER_OUTPUT):
        return send_file(
            TRACKER_OUTPUT,
            as_attachment=True,
            download_name="ACCT-108 Master Invoice Tracker 2026 - Updated.xlsx",
        )

    if os.path.exists(TRACKER_INPUT):
        return send_file(
            TRACKER_INPUT,
            as_attachment=True,
            download_name="ACCT-108 Master Invoice Tracker 2026.xlsx",
        )

    return jsonify({
        "error": "No tracker file found. Upload or run the workflow first."
    }), 404


@app.route("/api/clear/workspace", methods=["POST"])
def clear_workspace():
    import shutil

    deleted = {
        "invoices": 0,
        "client_files": 0,
        "output_files": 0,
        "input_files": 0,
        "media_plans": 0,
    }

    errors = []

    folders = [
        ("invoices", INVOICE_FOLDER, ".pdf"),
        ("client_files", CLIENTS_FOLDER, None),
        ("output_files", OUTPUT_FOLDER, None),
        ("input_files", INPUT_FOLDER, None),
        ("media_plans", MEDIA_PLANS_FOLDER, None),
    ]

    for key, folder, extension_filter in folders:
        if not os.path.exists(folder):
            continue

        for item in os.listdir(folder):
            if item == ".gitkeep":
                continue

            if extension_filter and not item.lower().endswith(extension_filter):
                continue

            item_path = os.path.join(folder, item)

            try:
                if os.path.isdir(item_path):
                    shutil.rmtree(item_path)
                else:
                    os.remove(item_path)

                deleted[key] += 1
            except Exception as e:
                errors.append(f"{item}: {e}")

    return jsonify({
        "message": (
            f"✓ Cleared {deleted['invoices']} invoice(s), "
            f"{deleted['client_files']} client folder(s), "
            f"{deleted['output_files']} output file(s), "
            f"{deleted['input_files']} input file(s), "
            f"{deleted['media_plans']} media plan(s)."
        ),
        "deleted": deleted,
        "errors": errors,
    }), 200


@app.route("/api/invoices", methods=["GET"])
def list_invoices():
    if not os.path.exists(INVOICE_FOLDER):
        return jsonify({"files": []})

    files = sorted([
        f for f in os.listdir(INVOICE_FOLDER)
        if f.lower().endswith(".pdf")
    ])

    return jsonify({"files": files})


@app.route("/api/invoices/delete", methods=["POST"])
def delete_invoice():
    data = request.get_json() or {}
    filenames = data.get("filenames", [])

    if isinstance(filenames, str):
        filenames = [filenames]

    if not filenames:
        return jsonify({"error": "No filenames provided"}), 400

    deleted = []
    errors = []

    for filename in filenames:
        safe_name = os.path.basename(secure_filename(filename))
        path = os.path.join(INVOICE_FOLDER, safe_name)

        if not os.path.exists(path):
            errors.append(f"{filename} not found")
            continue

        try:
            os.remove(path)
            deleted.append(safe_name)
        except Exception as e:
            errors.append(f"{filename}: {str(e)}")

    return jsonify({
        "message": f"Deleted {len(deleted)} file(s)" + (f", {len(errors)} error(s)" if errors else ""),
        "deleted": deleted,
        "errors": errors,
    }), 200 if deleted else 400


@app.route("/api/prisma/status", methods=["GET"])
def prisma_status():
    guide_loaded = False
    guide_rows = 0
    clients = []
    guide_exists = os.path.exists(BUYING_GUIDE_PATH)
    template_exists = os.path.exists(PRISMA_TEMPLATE)
    error = None

    if guide_exists:
        try:
            from buying_guide import BuyingGuide

            bg = BuyingGuide(BUYING_GUIDE_PATH)
            guide_loaded = True
            guide_rows = len(bg)
            clients = bg.clients()

        except Exception as e:
            error = str(e)

    gemini_state = gemini_status_payload()

    return jsonify({
        "guide_loaded": guide_loaded,
        "guide_exists": guide_exists,
        "guide_rows": guide_rows,
        "clients": clients,
        "template_exists": template_exists,
        "ready": guide_loaded and template_exists,
        "gemini": gemini_state["active"],
        "gemini_configured": gemini_state["configured"],
        "gemini_enabled": gemini_state["enabled"],
        "gemini_active": gemini_state["active"],
        "error": error,
    })


@app.route("/api/prisma/upload", methods=["POST"])
def prisma_upload():
    files = request.files.getlist("files")

    if not files:
        return jsonify({"error": "No files provided"}), 400

    saved = []
    errors = []

    for f in files:
        if not f or not allowed_plan_file(f.filename or ""):
            errors.append(f.filename)
            continue

        filename = secure_filename(f.filename)
        f.save(os.path.join(MEDIA_PLANS_FOLDER, filename))
        saved.append(filename)

    return jsonify({
        "message": f"✓ Uploaded {len(saved)} file(s)" + (f", {len(errors)} skipped" if errors else ""),
        "count": len(saved),
        "saved": saved,
        "errors": errors,
    })


@app.route("/api/prisma/plans", methods=["GET"])
def prisma_plans():
    if not os.path.exists(MEDIA_PLANS_FOLDER):
        return jsonify({"plans": []})

    try:
        from plan_parser import detect_client
    except Exception:
        detect_client = lambda _p: None

    plans = []

    for name in sorted(os.listdir(MEDIA_PLANS_FOLDER)):
        full = os.path.join(MEDIA_PLANS_FOLDER, name)

        if not os.path.isfile(full) or not allowed_plan_file(name):
            continue

        try:
            detected = detect_client(full)
        except Exception:
            detected = None

        plans.append({
            "filename": name,
            "size_kb": round(os.path.getsize(full) / 1024, 1),
            "detected_client": detected,
        })

    return jsonify({"plans": plans})


@app.route("/api/prisma/convert", methods=["POST"])
def prisma_convert():
    data = request.get_json() or {}

    filename = data.get("filename")
    forced_client = (data.get("client") or "").strip().upper()
    use_gemini = bool(data.get("use_gemini", False)) and gemini_available()
    skip_unmatched = bool(data.get("skip_unmatched_buying_guide", False))

    logs = []
    warnings = []
    diagnostics = {}

    def add_log(message):
        text = str(message)
        logs.append(text)

    if not filename:
        return jsonify({
            "ok": False,
            "error": "filename required",
            "logs": logs,
            "warnings": warnings,
            "diagnostics": diagnostics,
        }), 400

    safe_filename = secure_filename(filename)
    media_plan_path = os.path.join(MEDIA_PLANS_FOLDER, safe_filename)

    if not os.path.exists(media_plan_path):
        return jsonify({
            "ok": False,
            "error": f"file not found: {safe_filename}",
            "logs": logs,
            "warnings": warnings,
            "diagnostics": diagnostics,
        }), 404

    if not os.path.exists(BUYING_GUIDE_PATH):
        return jsonify({
            "ok": False,
            "error": "Buying Guide missing — upload it via the Reference Files card",
            "logs": logs,
            "warnings": warnings,
            "diagnostics": diagnostics,
        }), 400

    os.makedirs(PRISMA_OUTPUT_DIR, exist_ok=True)

    base_name = os.path.splitext(safe_filename)[0]

    try:
        from plan_parser import detect_client
        from run_media import run_media_plan_to_prisma

        detected_client = detect_client(media_plan_path)
        client = (forced_client or detected_client or "GU").upper()

        output_filename = f"{client}_{base_name}_prisma_import.xlsx"
        output_path = os.path.join(PRISMA_OUTPUT_DIR, output_filename)

        diagnostics.update({
            "filename": safe_filename,
            "detected_client": detected_client,
            "forced_client": forced_client or None,
            "client": client,
            "use_gemini": use_gemini,
            "gemini_configured": gemini_configured(),
            "gemini_enabled": get_gemini_enabled(),
            "skip_unmatched_buying_guide": skip_unmatched,
        })

        result = run_media_plan_to_prisma(
            media_plan_path=media_plan_path,
            buying_guide_path=BUYING_GUIDE_PATH,
            output_path=output_path,
            client=client,
            debug=False,
            use_gemini=use_gemini,
            skip_unmatched_buying_guide=skip_unmatched,
            generate_buying_guide_gap_report=True,
            verbose=False,
        )

        logs.extend(result.get("logs", []))
        warnings.extend(result.get("warnings", []))
        diagnostics.update(result.get("diagnostics", {}))

        final_output_file = result.get("output_file") or output_filename
        preview_records = result.get("preview", [])

        matched_count = 0
        unmatched = []

        for record in preview_records:
            status = str(record.get("status", ""))

            if status.lower().startswith("matched"):
                matched_count += 1
            else:
                unmatched.append(record.get("partner"))

        gap_report_path = diagnostics.get("buying_guide_gap_report_path", "")
        gap_report_file = os.path.basename(gap_report_path) if gap_report_path else ""

        response = {
            "ok": True,
            "client": client,
            "detected_client": detected_client,
            "forced_client": forced_client or None,
            "adapter": "run_media.py",
            "placement_count": diagnostics.get("consolidated_rows"),
            "matched_count": matched_count,
            "unmatched_channels": unmatched,
            "output_file": final_output_file,
            "download_url": f"/api/prisma/download/{final_output_file}",
            "preview": preview_records,
            "logs": logs,
            "warnings": warnings,
            "diagnostics": diagnostics,
        }

        if gap_report_file:
            response["gap_report_file"] = gap_report_file
            response["gap_report_download_url"] = f"/api/prisma/download/{gap_report_file}"

        return jsonify(response)

    except Exception as e:
        error_message = str(e)
        add_log(f"Conversion failed: {error_message}")

        try:
            from plan_parser import detect_client
            detected_client = detect_client(media_plan_path)
        except Exception:
            detected_client = None

        client = (forced_client or detected_client or "GU").upper()

        diagnostics.update({
            "filename": safe_filename,
            "client": client,
            "detected_client": detected_client,
            "forced_client": forced_client or None,
            "use_gemini": use_gemini,
            "gemini_configured": gemini_configured(),
            "gemini_enabled": get_gemini_enabled(),
            "skip_unmatched_buying_guide": skip_unmatched,
        })

        return jsonify({
            "ok": False,
            "error": error_message,
            "message": error_message,
            "filename": safe_filename,
            "client": client,
            "detected_client": detected_client,
            "forced_client": forced_client or None,
            "logs": logs,
            "warnings": warnings,
            "diagnostics": diagnostics,
        }), 500


@app.route("/api/prisma/download/<path:name>", methods=["GET"])
def prisma_download(name):
    safe_name = secure_filename(name)

    allowed_suffixes = (
        "_prisma_import.xlsx",
        "_buying_guide_gap_report.xlsx",
    )

    if not safe_name.endswith(allowed_suffixes):
        return jsonify({
            "error": "Unsupported Prisma download file."
        }), 400

    full = os.path.join(PRISMA_OUTPUT_DIR, safe_name)

    if not os.path.exists(full):
        return jsonify({
            "error": "File not found"
        }), 404

    return send_file(
        full,
        as_attachment=True,
        download_name=safe_name,
    )


@app.route("/api/prisma/plans/delete", methods=["POST"])
def prisma_delete_plans():
    data = request.get_json() or {}
    filenames = data.get("filenames", [])

    if isinstance(filenames, str):
        filenames = [filenames]

    if not filenames:
        return jsonify({"error": "No filenames provided"}), 400

    deleted = []
    errors = []

    for filename in filenames:
        safe = os.path.basename(secure_filename(filename))
        path = os.path.join(MEDIA_PLANS_FOLDER, safe)

        if not os.path.exists(path):
            errors.append(f"{filename} not found")
            continue

        try:
            os.remove(path)
            deleted.append(safe)
        except Exception as e:
            errors.append(f"{filename}: {str(e)}")

    return jsonify({
        "message": f"Deleted {len(deleted)} file(s)" + (f", {len(errors)} error(s)" if errors else ""),
        "deleted": deleted,
        "errors": errors,
    }), 200 if deleted else 400


@app.route("/api/prisma/files/delete", methods=["POST"])
def delete_prisma_reference_files():
    data = request.get_json() or {}
    targets = data.get("targets", [])

    if isinstance(targets, str):
        targets = [targets]

    deleted = []
    errors = []

    if "buying_guide" in targets:
        if os.path.exists(BUYING_GUIDE_PATH):
            try:
                os.remove(BUYING_GUIDE_PATH)
                deleted.append("buying_guide")
            except Exception as e:
                errors.append(f"buying_guide: {e}")
        else:
            errors.append("buying_guide not found")

    if "template" in targets:
        if os.path.exists(PRISMA_TEMPLATE):
            try:
                os.remove(PRISMA_TEMPLATE)
                deleted.append("template")
            except Exception as e:
                errors.append(f"template: {e}")
        else:
            errors.append("template not found")

    return jsonify({
        "message": f"Deleted: {', '.join(deleted) or 'nothing'}",
        "deleted": deleted,
        "errors": errors,
    }), 200 if deleted else 400


@app.route("/api/run/sort")
def run_sort():
    return run_and_stream("invoice_sorter.py")


@app.route("/api/run/extract")
def run_extract():
    return run_and_stream("invoice_extractor.py")


@app.route("/api/run/workflow")
def run_workflow():
    return run_and_stream("run_workflow.py")


@app.route("/api/drive/sync", methods=["POST"])
def drive_sync():
    if not os.path.exists(CREDENTIALS):
        return jsonify({
            "status": "not_configured",
            "message": "[WARN] credentials.json not found. Upload it via the dashboard first.",
        }), 501

    if not os.path.exists(DRIVE_CONFIG):
        return jsonify({
            "status": "not_configured",
            "message": "[WARN] Drive folder ID not set. Add it in the Google Drive Sync card.",
        }), 501

    return run_and_stream("drive_sync.py")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)