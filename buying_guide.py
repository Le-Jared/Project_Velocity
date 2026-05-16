"""
Loads and indexes the ACCT 108 BuyingGuide.xlsx reference file.
Provides client-filtered supplier-code lookups for Prisma conversion.
"""
from __future__ import annotations
import re
from pathlib import Path
from typing import Optional

import openpyxl


SHEET_NAME = "Buying guide 2"


class BuyingGuide:
    """In-memory index of the Buying Guide for fast supplier lookups."""

    # ── Channel alias map ────────────────────────────────────────────────
    # Maps normalized channel names (from media plans) → normalized booking
    # keys (from the Buying Guide "Placement booking type" column).
    # Add new aliases here whenever a media plan uses a non-standard name.
    CHANNEL_ALIASES: dict[str, str] = {
        "asa":        "apple search",   # Apple Search Ads
        "google uac": "google display", # Universal App Campaign
        "uac":        "google display", # UAC shorthand (SkillIgnition)
        "meta":       "facebook",       # Meta = Facebook supplier
    }

    def __init__(self, path: Path | str):
        self.path = Path(path)
        self.rows: list[dict] = []
        if not self.path.exists():
            raise FileNotFoundError(f"Buying Guide not found at {self.path}")
        self._load()

    # ── Loading ──────────────────────────────────────────────────────────

    def _load(self) -> None:
        wb = openpyxl.load_workbook(self.path, data_only=True, read_only=True)

        # Sheet name may have trailing spaces or case differences — be lenient
        target_sheet = None
        for sn in wb.sheetnames:
            if sn.strip().lower() == SHEET_NAME.lower():
                target_sheet = sn
                break
        if not target_sheet:
            wb.close()
            raise ValueError(f"Sheet '{SHEET_NAME}' not found in {self.path.name}")

        ws = wb[target_sheet]
        rows_iter = ws.iter_rows(values_only=True)

        try:
            header_row = next(rows_iter)
        except StopIteration:
            wb.close()
            return

        headers = [str(c).strip() if c is not None else "" for c in header_row]

        for raw in rows_iter:
            if not raw or not any(raw):
                continue
            row = dict(zip(headers, raw))

            client_cell = str(row.get("Clients that uses these respectively") or "").strip()
            if not client_cell or client_cell.upper() == "NA":
                row["_clients"] = []
            else:
                row["_clients"] = [
                    c.strip().upper()
                    for c in re.split(r"[,/]| and ", client_cell)
                    if c.strip()
                ]

            booking = str(row.get("Placement booking type") or "")
            row["_booking"]     = booking
            row["_booking_key"] = self._normalize(booking.replace("- Client Paying Supplier", ""))
            row["_is_cps"]      = "client paying supplier" in booking.lower()
            self.rows.append(row)

        wb.close()

    # ── Helpers ──────────────────────────────────────────────────────────

    @staticmethod
    def _normalize(text: str) -> str:
        """Lowercase + collapse non-alphanumeric for fuzzy matching."""
        t = (text or "").lower()
        t = re.sub(r"[^a-z0-9]+", " ", t).strip()
        return t

    # ── Internal lookup (single CPS flag) ────────────────────────────────

    def _lookup_with_cps(
        self,
        client_u: str,
        channel_norm: str,
        currency: Optional[str],
        client_paying_supplier: bool,
    ) -> Optional[dict]:
        candidates = []
        for row in self.rows:
            if client_u and client_u not in row["_clients"]:
                continue
            if row["_is_cps"] != client_paying_supplier:
                continue
            if currency and str(row.get("Currency") or "").upper() != currency.upper():
                continue

            booking_key = row["_booking_key"]
            if not booking_key:
                continue

            # Score: exact > contains > token overlap
            if booking_key == channel_norm:
                score = 100
            elif channel_norm in booking_key or booking_key in channel_norm:
                score = 50 - abs(len(booking_key) - len(channel_norm))
            else:
                ch_tokens = set(channel_norm.split())
                bk_tokens = set(booking_key.split())
                overlap   = len(ch_tokens & bk_tokens)
                if overlap == 0:
                    continue
                score = overlap * 10

            candidates.append((score, row))

        if not candidates:
            return None
        candidates.sort(key=lambda x: x[0], reverse=True)
        return candidates[0][1]

    # ── Public API ───────────────────────────────────────────────────────

    def lookup(
        self,
        client: str,
        channel: str,
        currency: Optional[str] = None,
        client_paying_supplier: bool = False,
    ) -> Optional[dict]:
        """
        Find best-matching guide row for a given client + channel.

        Resolution order:
          1. Apply channel alias (e.g. 'uac' → 'google display').
          2. Try exact client_paying_supplier flag as passed.
          3. If no match, automatically retry with the opposite CPS flag
             so callers don't need to know whether a client uses CPS or not.
        """
        client_u     = (client or "").upper()
        channel_norm = self._normalize(channel)
        if not channel_norm:
            return None

        # Step 1 — apply alias
        channel_norm = self.CHANNEL_ALIASES.get(channel_norm, channel_norm)

        # Step 2 — try as requested
        result = self._lookup_with_cps(client_u, channel_norm, currency, client_paying_supplier)
        if result:
            return result

        # Step 3 — fallback: retry with opposite CPS flag
        return self._lookup_with_cps(client_u, channel_norm, currency, not client_paying_supplier)

    def clients(self) -> list[str]:
        """Return sorted unique client codes seen in the guide."""
        seen = set()
        for r in self.rows:
            for c in r["_clients"]:
                seen.add(c)
        return sorted(seen)

    def __len__(self) -> int:
        return len(self.rows)