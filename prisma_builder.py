from pathlib import Path
import copy
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Alignment, Protection
from openpyxl.utils import get_column_letter


PRISMA_COLUMNS = [
    "Notes",
    "Site Name/Supplier",
    "PACKAGE/PLACEMENT TYPE",
    "Buy Type",
    "Buy Category",
    "Booking Category",
    "Package name",
    "Placement Name",
    "Unit Dimensions",
    "Positioning",
    "Cost Method",
    "Unit Type",
    "Unit Rate",
    "Planned unit amount",
    "Gross/Planned Cost",
    "Flight Start",
    "Flight End",
    "Programmatic Provider",
    "Market place",
    "Social network",
    "Strategy",
]


REQUIRED_TEXT_COLUMNS = [
    "Site Name/Supplier",
    "PACKAGE/PLACEMENT TYPE",
    "Buy Type",
    "Buy Category",
    "Placement Name",
    "Cost Method",
    "Unit Type",
    "Flight Start",
    "Flight End",
]


REQUIRED_NUMERIC_COLUMNS = [
    "Unit Rate",
    "Planned unit amount",
    "Gross/Planned Cost",
]


COLUMN_WIDTHS = {
    "Placement Name": 30,
    "Site Name/Supplier": 30,
    "Flight Start": 15,
    "Flight End": 15,
    "Gross/Planned Cost": 18,
    "Planned unit amount": 18,
    "Unit Rate": 12,
}


TEMPLATE_SHEET_NAME = "Digital import sheet ALL TYPES"


def _is_blank(value):
    if value is None:
        return True

    try:
        if pd.isna(value):
            return True
    except Exception:
        pass

    return str(value).strip().lower() in {"", "nan", "none", "na", "n/a"}


def _first_non_empty(*values):
    for value in values:
        if not _is_blank(value):
            return value

    return ""


def _safe_float(value, default=0.0):
    if _is_blank(value):
        return default

    try:
        text = str(value).replace(",", "").replace("$", "").strip()
        return float(text)
    except Exception:
        return default


def _format_excel_date(value):
    if _is_blank(value):
        return ""

    parsed = pd.to_datetime(value, errors="coerce")

    if pd.isna(parsed):
        return ""

    return parsed.to_pydatetime()


def _format_date_for_validation(value):
    parsed = pd.to_datetime(value, errors="coerce")

    if pd.isna(parsed):
        return ""

    return parsed.strftime("%m/%d/%Y")


def _is_reasonable_prisma_date(value):
    parsed = pd.to_datetime(value, errors="coerce")

    if pd.isna(parsed):
        return False

    return 2020 <= parsed.year <= 2035


def _resolve_flight_start(row):
    return _first_non_empty(
        row.get("flight_start"),
        row.get("start_date"),
        row.get("campaign_start"),
        row.get("Campaign Start"),
        row.get("Start Date"),
    )


def _resolve_flight_end(row):
    return _first_non_empty(
        row.get("flight_end"),
        row.get("end_date"),
        row.get("campaign_end"),
        row.get("Campaign End"),
        row.get("End Date"),
    )


def _row_skip_reasons(row):
    planned_amount = _safe_float(row.get("planned_amount", 0))
    planned_units = _safe_float(row.get("planned_units", 0))
    flight_start = _resolve_flight_start(row)
    flight_end = _resolve_flight_end(row)

    reasons = []

    if planned_amount <= 0:
        reasons.append("planned_amount <= 0")

    if planned_units <= 0:
        reasons.append("planned_units <= 0")

    if _is_blank(flight_start):
        reasons.append("missing flight_start")

    if _is_blank(flight_end):
        reasons.append("missing flight_end")

    if not _is_blank(flight_start) and not _is_reasonable_prisma_date(flight_start):
        reasons.append(f"invalid flight_start: {_format_date_for_validation(flight_start)}")

    if not _is_blank(flight_end) and not _is_reasonable_prisma_date(flight_end):
        reasons.append(f"invalid flight_end: {_format_date_for_validation(flight_end)}")

    return reasons


def _build_prisma_row(row):
    cost_method = row.get("guide_cost_method") or row.get("cost_method") or "CPM"
    unit_type = row.get("guide_unit_type") or row.get("unit_type") or "Impressions"

    flight_start = _resolve_flight_start(row)
    flight_end = _resolve_flight_end(row)

    return {
        "Notes": "Direct placement",
        "Site Name/Supplier": row.get("supplier_name", ""),
        "PACKAGE/PLACEMENT TYPE": "Standalone",
        "Buy Type": row.get("buy_type", "Display"),
        "Buy Category": row.get("buy_category", "Mobile"),
        "Booking Category": "Standard",
        "Package name": "",
        "Placement Name": row.get("placement_name", ""),
        "Unit Dimensions": row.get("ad_size", "1 x 1"),
        "Positioning": row.get("positioning", "Other"),
        "Cost Method": cost_method,
        "Unit Type": unit_type,
        "Unit Rate": _safe_float(row.get("unit_rate", 0)),
        "Planned unit amount": int(round(_safe_float(row.get("planned_units", 0)), 0)),
        "Gross/Planned Cost": round(_safe_float(row.get("planned_amount", 0)), 2),
        "Flight Start": _format_excel_date(flight_start),
        "Flight End": _format_excel_date(flight_end),
        "Programmatic Provider": "",
        "Market place": "",
        "Social network": "",
        "Strategy": "",
    }


def _format_skipped_rows(skipped, limit=10):
    if not skipped:
        return ""

    parts = [
        f"Row {item['row']} partner='{item['partner']}' placement='{item['placement_name']}': {item['reasons']}"
        for item in skipped[:limit]
    ]

    remaining = len(skipped) - limit

    if remaining > 0:
        parts.append(f"... plus {remaining} more skipped row(s)")

    return "; ".join(parts)


def build_prisma_dataframe(enriched_df):
    if enriched_df.empty:
        raise ValueError("Cannot build Prisma dataframe from empty data.")

    rows = []
    skipped = []

    for idx, row in enriched_df.iterrows():
        reasons = _row_skip_reasons(row)

        if reasons:
            skipped.append(
                {
                    "row": int(idx) + 2,
                    "partner": row.get("partner", ""),
                    "placement_name": row.get("placement_name", ""),
                    "reasons": ", ".join(reasons),
                }
            )
            continue

        rows.append(_build_prisma_row(row))

    if not rows:
        raise ValueError(
            "No valid Prisma rows after filtering. "
            "Rows must have valid flight dates, planned units, and gross/planned cost greater than 0. "
            f"Skipped rows: {_format_skipped_rows(skipped)}"
        )

    return pd.DataFrame(rows, columns=PRISMA_COLUMNS)


def validate_prisma_dataframe(prisma_df):
    errors = []
    missing_columns = [col for col in PRISMA_COLUMNS if col not in prisma_df.columns]

    if missing_columns:
        return [f"Missing Prisma column: {col}" for col in missing_columns]

    for idx, row in prisma_df.iterrows():
        excel_row = idx + 2

        for col in REQUIRED_TEXT_COLUMNS:
            value = row.get(col)

            if _is_blank(value):
                errors.append(f"Row {excel_row}: Missing required field '{col}'")

        for col in ["Flight Start", "Flight End"]:
            value = row.get(col)

            if not _is_reasonable_prisma_date(value):
                errors.append(
                    f"Row {excel_row}: Invalid or unreasonable '{col}': {_format_date_for_validation(value)}"
                )

        for col in REQUIRED_NUMERIC_COLUMNS:
            value = row.get(col)
            numeric_value = _safe_float(value, default=None)

            if numeric_value is None:
                errors.append(f"Row {excel_row}: Field '{col}' is not numeric")
                continue

            if numeric_value < 0:
                errors.append(f"Row {excel_row}: Field '{col}' cannot be negative")

        gross_cost = _safe_float(row.get("Gross/Planned Cost", 0))
        planned_units = _safe_float(row.get("Planned unit amount", 0))

        if gross_cost <= 0:
            errors.append(f"Row {excel_row}: Gross/Planned Cost must be greater than 0")

        if planned_units <= 0:
            errors.append(f"Row {excel_row}: Planned unit amount must be greater than 0")

    return errors


def _normalize_header(value):
    return str(value or "").replace("\n", " ").strip().lower()


def _find_template_header_row(ws):
    target = {_normalize_header(col) for col in PRISMA_COLUMNS}

    for row_idx in range(1, min(ws.max_row, 30) + 1):
        values = [_normalize_header(ws.cell(row=row_idx, column=col_idx).value) for col_idx in range(1, ws.max_column + 1)]
        hits = sum(1 for value in values if value in target)

        if hits >= 10:
            return row_idx

    raise ValueError("Could not locate Prisma template header row.")


def _template_column_map(ws, header_row):
    mapping = {}

    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col_idx).value
        normalized = _normalize_header(value)

        for prisma_col in PRISMA_COLUMNS:
            if normalized == _normalize_header(prisma_col):
                mapping[prisma_col] = col_idx
                break

    missing = [col for col in PRISMA_COLUMNS if col not in mapping]

    if missing:
        raise ValueError("Template is missing required Prisma columns: " + ", ".join(missing))

    return mapping


def _copy_row_style(ws, source_row, target_row):
    for col_idx in range(1, ws.max_column + 1):
        source = ws.cell(row=source_row, column=col_idx)
        target = ws.cell(row=target_row, column=col_idx)

        if source.has_style:
            target._style = copy.copy(source._style)

        if source.number_format:
            target.number_format = source.number_format

        if source.font:
            target.font = copy.copy(source.font)

        if source.fill:
            target.fill = copy.copy(source.fill)

        if source.border:
            target.border = copy.copy(source.border)

        if source.alignment:
            target.alignment = copy.copy(source.alignment)

        if source.protection:
            target.protection = copy.copy(source.protection)


def _clear_existing_data_rows(ws, header_row):
    first_data_row = header_row + 1

    for row_idx in range(first_data_row, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).value = None


def _write_dataframe_to_template(ws, prisma_df, header_row, column_map):
    first_data_row = header_row + 1
    style_source_row = first_data_row

    for idx, row in prisma_df.iterrows():
        excel_row = first_data_row + idx

        if excel_row != style_source_row:
            _copy_row_style(ws, style_source_row, excel_row)

        for col_name in PRISMA_COLUMNS:
            cell = ws.cell(row=excel_row, column=column_map[col_name])
            value = row.get(col_name, "")

            if col_name in {"Flight Start", "Flight End"} and not _is_blank(value):
                cell.value = pd.to_datetime(value).to_pydatetime()
                cell.number_format = "mm/dd/yyyy"
            elif col_name == "Gross/Planned Cost":
                cell.value = _safe_float(value)
                cell.number_format = "$#,##0.00"
            elif col_name == "Planned unit amount":
                cell.value = int(round(_safe_float(value), 0))
                cell.number_format = "#,##0"
            elif col_name == "Unit Rate":
                cell.value = round(_safe_float(value), 2)
                cell.number_format = "0.00"
            else:
                cell.value = value


def _apply_template_widths(ws, column_map):
    for col_name, width in COLUMN_WIDTHS.items():
        if col_name in column_map:
            ws.column_dimensions[get_column_letter(column_map[col_name])].width = width


def export_prisma_import(enriched_df, output_path, template_path=None):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    prisma_df = build_prisma_dataframe(enriched_df)
    validation_errors = validate_prisma_dataframe(prisma_df)

    if validation_errors:
        raise ValueError("Prisma export validation failed:\n" + "\n".join(validation_errors))

    if template_path:
        return export_prisma_import_from_template(prisma_df, output_path, template_path)

    return export_prisma_import_plain(prisma_df, output_path)


def export_prisma_import_from_template(prisma_df, output_path, template_path):
    template_path = Path(template_path)

    if not template_path.exists():
        raise FileNotFoundError(f"Prisma template not found: {template_path}")

    wb = load_workbook(template_path)
    ws = wb[TEMPLATE_SHEET_NAME] if TEMPLATE_SHEET_NAME in wb.sheetnames else wb.active

    header_row = _find_template_header_row(ws)
    column_map = _template_column_map(ws, header_row)

    _clear_existing_data_rows(ws, header_row)
    _write_dataframe_to_template(ws, prisma_df, header_row, column_map)
    _apply_template_widths(ws, column_map)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{header_row + len(prisma_df)}"

    wb.save(output_path)

    return output_path


def export_prisma_import_plain(prisma_df, output_path):
    with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
        prisma_df.to_excel(
            writer,
            sheet_name=TEMPLATE_SHEET_NAME,
            index=False,
            startrow=0,
        )

        workbook = writer.book
        worksheet = writer.sheets[TEMPLATE_SHEET_NAME]

        header_format = workbook.add_format(
            {
                "bold": True,
                "bg_color": "#D9EAF7",
                "border": 1,
                "text_wrap": True,
                "valign": "top",
            }
        )

        formats = {
            "Gross/Planned Cost": workbook.add_format({"num_format": "$#,##0.00"}),
            "Planned unit amount": workbook.add_format({"num_format": "#,##0"}),
            "Unit Rate": workbook.add_format({"num_format": "0.00"}),
            "Flight Start": workbook.add_format({"num_format": "mm/dd/yyyy"}),
            "Flight End": workbook.add_format({"num_format": "mm/dd/yyyy"}),
        }

        for col_idx, col_name in enumerate(prisma_df.columns):
            worksheet.write(0, col_idx, col_name, header_format)

            width = COLUMN_WIDTHS.get(col_name, 18)
            cell_format = formats.get(col_name)

            if cell_format:
                worksheet.set_column(col_idx, col_idx, width, cell_format)
            else:
                worksheet.set_column(col_idx, col_idx, width)

        worksheet.freeze_panes(1, 0)
        worksheet.autofilter(0, 0, len(prisma_df), len(prisma_df.columns) - 1)

    return output_path


def export_debug_files(raw_df, normalized_df, consolidated_df, enriched_df, output_dir):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    files = {
        "01_raw_parsed.csv": raw_df,
        "02_normalized.csv": normalized_df,
        "03_consolidated.csv": consolidated_df,
        "04_enriched.csv": enriched_df,
    }

    for filename, df in files.items():
        df.to_csv(output_dir / filename, index=False)
