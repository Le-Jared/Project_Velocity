import os
import re
import sys
import warnings
from time import perf_counter
from datetime import datetime
from calendar import month_name

import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from gemini_fallback import enrich_extraction, is_available as gemini_available


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

warnings.filterwarnings("ignore")


INPUT_FOLDER = "./Input"
OUTPUT_FOLDER = "./Output"
TRACKER_PATH = os.path.join(INPUT_FOLDER, "ACCT-108 Master Invoice Tracker 2026.xlsx")
OUTPUT_PATH = os.path.join(OUTPUT_FOLDER, "ACCT-108 Master Invoice Tracker 2026 - Updated.xlsx")
INVOICE_FOLDER = "./invoices"

YEAR = 2026
SEP = "=" * 72
SUB_SEP = "-" * 72
WHITE = "FFFFFF"
DARK = "1A1A1A"

COLOURS = {
    "adsjoy": {"header_bg": "1F3864", "row_alt": "DCE6F1"},
    "apple": {"header_bg": "1C1C1E", "row_alt": "E8E8E8"},
    "google": {"header_bg": "1A73E8", "row_alt": "E8F0FE"},
    "meta": {"header_bg": "1877F2", "row_alt": "E7F0FD"},
}

CURRENCY_COLOURS = {
    "IDR": "E8F5E9",
    "MYR": "FFF8E1",
    "SGD": "E3F2FD",
    "USD": "F3E5F5",
    "GBP": "ECEFF1",
}

CURRENCY_HEADER = {
    "IDR": "388E3C",
    "MYR": "F57F17",
    "SGD": "1565C0",
    "USD": "6A1B9A",
    "GBP": "455A64",
}

META_ENTITY_MARKET = {
    "FACEBOOK SINGAPORE": "SG",
    "FACEBOOK UK": "GB",
    "FACEBOOK IRELAND": "IE",
    "FACEBOOK NETHERLANDS": "NL",
    "FACEBOOK AUSTRALIA": "AU",
    "FACEBOOK THAILAND": "TH",
    "FACEBOOK MALAYSIA": "MY",
    "FACEBOOK INDONESIA": "ID",
    "FACEBOOK PHILIPPINES": "PH",
}

META_ACCOUNT_CLIENT = {
    "10472231667355": "BHC",
    "10572631630900": "LGL",
}

SHEET_NAMES = {
    "adsjoy": "Adsjoy",
    "apple": "Apple (ASA)",
    "google": "Google",
    "meta": "Meta (facebook)",
}

SUPPLIER_LABELS = {
    "adsjoy": "AdsJoy",
    "apple": "Apple (ASA)",
    "google": "Google",
    "meta": "Meta (Facebook)",
}

STANDARD_SUPPLIER_NAMES = {
    "adsjoy": "AdsJoy",
    "AdsJoy": "AdsJoy",
    "Adsjoy": "AdsJoy",
    "apple": "Apple (ASA)",
    "Apple ( ASA )": "Apple (ASA)",
    "Apple (ASA)": "Apple (ASA)",
    "Apple ASA": "Apple (ASA)",
    "Apple Search Ads": "Apple (ASA)",
    "google": "Google",
    "Google": "Google",
    "meta": "Meta (Facebook)",
    "Meta": "Meta (Facebook)",
    "Meta (facebook)": "Meta (Facebook)",
    "Meta (Facebook)": "Meta (Facebook)",
    "Facebook": "Meta (Facebook)",
    "Meta Facebook": "Meta (Facebook)",
}

CLIENT_CANONICAL = {
    "AMA": "Ama",
    "AMA PROSPER": "Ama",
    "FJPH": "FJPH",
    "AFID": "AFID",
    "BMYFF": "BMYFF",
    "LGL": "LGL",
    "BHC": "BHC",
    "MF": "MF",
    "PD": "PD",
    "OG": "OG",
    "GU": "GU",
    "GT": "GT",
}

STANDARD_COLUMNS = {
    "Adsjoy": [
        "Year", "Supplier Name", "Month of Service", "Month of Billing",
        "Client", "Invoice number", "Campaign", "Campaign ID", "Currency", "Amount",
    ],
    "Apple (ASA)": [
        "Year", "Supplier Name", "Client", "Month of Service", "Month of Billing",
        "Invoice number", "Campaign", "Campaign ID", "Currency", "Amount", "Market",
    ],
    "Google": [
        "Year", "Supplier Name", "Ad Account Name", "Month of Service", "Month of Billing",
        "Client", "Invoice number", "Campaign", "Campaign ID", "Currency", "Amount",
    ],
    "Meta (facebook)": [
        "Year", "Supplier Name", "Ad Account ID", "Month of Service", "Month of Billing",
        "Client", "Invoice number", "Campaign", "Campaign ID", "Currency", "Amount", "Market",
    ],
}

MONTH_ABBR = {m[:3].lower(): m for m in month_name if m}


def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def duration_str(seconds):
    if seconds < 60:
        return f"{seconds:.1f}s"
    return f"{int(seconds // 60)}m {seconds % 60:.1f}s"


def log(level, msg="", indent=0):
    print(f"{' ' * indent}[{level}] {msg}" if msg else "")


def print_header():
    print(SEP)
    print("ACCT-108 Invoice Extractor")
    print("Suppliers : AdsJoy | Apple ASA | Google | Meta")
    print("Mode      : PDF-content-first + Gemini fallback")
    print(f"Started   : {now_str()}")
    print(SEP)

    if gemini_available():
        log("GEMINI", "Fallback active — unknown fields resolved via AI")
        log("GEMINI", "Native PDF mode auto-enabled for low-text invoices")
    else:
        log("GEMINI", "Not configured — regex-only mode")


def clean_amount(val):
    if val is None:
        return None

    val = str(val).strip()

    if val == "" or val.lower() in {"nan", "none", "null", "nat"}:
        return None

    negative = val.startswith("(") and val.endswith(")")
    cleaned = re.sub(r"[^\d.-]", "", val.replace(",", "").replace(" ", ""))

    if cleaned in {"", ".", "-", "-."}:
        return None

    try:
        amount = float(cleaned)
        return -abs(amount) if negative else amount
    except ValueError:
        return None


def is_reasonable_amount(amount, currency):
    amount = clean_amount(amount)

    if amount is None:
        return False

    limits = {
        "USD": 5_000_000,
        "SGD": 5_000_000,
        "MYR": 20_000_000,
        "GBP": 5_000_000,
        "IDR": 100_000_000_000,
    }

    currency = str(currency or "").strip().upper()
    return abs(amount) <= limits.get(currency, 10_000_000)


def first_match(patterns, text, flags=re.I, group=1, default=""):
    if isinstance(patterns, str):
        patterns = [patterns]

    for pattern in patterns:
        match = re.search(pattern, text, flags)
        if match:
            return match.group(group).strip()

    return default


def normalize_month(mon, year):
    if not mon or not year:
        return ""

    year = str(year)
    if len(year) == 2:
        year = f"20{year}"

    return f"{MONTH_ABBR.get(mon[:3].lower(), mon.capitalize())} {year}"


def normalize_month_value(value):
    if pd.isna(value) or value == "":
        return ""

    value = str(value).strip()

    if value.lower() in {"nan", "none", "null", "nat"}:
        return ""

    m = re.match(r"^([A-Za-z]+)\s+(20\d{2})$", value)
    if m:
        return normalize_month(m.group(1), m.group(2))

    m = re.match(r"^([A-Za-z]{3})-?(\d{2})$", value)
    if m:
        return normalize_month(m.group(1), m.group(2))

    try:
        dt = pd.to_datetime(value, errors="coerce")
        if pd.notna(dt):
            return dt.strftime("%B %Y")
    except Exception:
        pass

    return value


def billing_period_to_str(raw):
    if not raw:
        return ""

    raw = raw.strip()

    m = re.match(r"([A-Za-z]{3})-?(\d{2})$", raw)
    if m:
        return normalize_month(m.group(1), m.group(2))

    m = re.match(r"([A-Za-z]+)\s+(20\d{2})$", raw)
    if m:
        return normalize_month(m.group(1), m.group(2))

    return normalize_month_value(raw)


def extract_month_from_text_or_filename(text, filename):
    patterns = [
        r"([A-Za-z]{3})'(\d{2})",
        r"Summary\s+for\s+\d{1,2}\s+([A-Za-z]+)\s+(\d{4})",
        r"Billing\s*Period\s*[:\s]*\d{1,2}\s+([A-Za-z]+)\s+(\d{4})",
        r"Billing\s*Period[:\s]*([A-Za-z]{3}-\d{2})",
        r"([A-Za-z]{3,})\s+(20\d{2})",
    ]

    for pattern in patterns:
        m = re.search(pattern, text, re.I)
        if m:
            return billing_period_to_str(m.group(1)) if len(m.groups()) == 1 else normalize_month(m.group(1), m.group(2))

    m = re.search(r"_([A-Za-z]{3})_(\d{2})_", filename, re.I)
    return normalize_month(m.group(1), m.group(2)) if m else ""


def extract_currency(text, default="USD"):
    match = re.search(r"\b(USD|SGD|MYR|IDR|AUD|GBP|PHP|INR)\b", text, re.I)
    return match.group(1).upper() if match else default


def normalize_supplier_name(value):
    if pd.isna(value):
        return ""

    value = str(value).strip()
    return STANDARD_SUPPLIER_NAMES.get(value, value)


def normalize_client(value):
    if pd.isna(value):
        return ""

    value = str(value).strip()

    if value.lower() in {"nan", "none", "null", "nat"}:
        return ""

    return CLIENT_CANONICAL.get(value.upper(), value)


def get_invoice_col(df):
    if df is None or df.empty:
        return None

    return next(
        (c for c in df.columns if "invoice" in str(c).lower() and "number" in str(c).lower()),
        None,
    )


def invoice_exists(tracker_df, invoice_number):
    if tracker_df is None or tracker_df.empty or not invoice_number:
        return False

    inv_col = get_invoice_col(tracker_df)
    if not inv_col:
        return False

    return str(invoice_number).strip() in tracker_df[inv_col].astype(str).str.strip().values


def invoice_has_campaign_rows(tracker_df, invoice_number):
    if tracker_df is None or tracker_df.empty or not invoice_number:
        return False

    inv_col = get_invoice_col(tracker_df)

    if not inv_col or "Campaign" not in tracker_df.columns:
        return False

    sub = tracker_df[
        tracker_df[inv_col].astype(str).str.strip().eq(str(invoice_number).strip())
    ]

    return not sub.empty and sub["Campaign"].fillna("").astype(str).str.strip().ne("").any()


def clean_text_columns(df):
    if df.empty:
        return df

    df = df.copy()

    for col in df.columns:
        if df[col].dtype == "object":
            df[col] = (
                df[col]
                .astype(str)
                .replace({"nan": "", "None": "", "NaT": "", "nat": ""})
                .str.strip()
            )

    return df


def remove_duplicate_columns(df):
    if df.empty:
        return df

    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    df = df.loc[:, ~df.columns.duplicated()]

    if "Month" in df.columns and "Year" in df.columns:
        df = df.drop(columns=["Month"], errors="ignore")

    currency_cols = [c for c in df.columns if c.strip().lower() == "currency"]

    if len(currency_cols) > 1:
        base = currency_cols[0]
        for c in currency_cols[1:]:
            df[base] = df[base].replace("", pd.NA).fillna(df[c])
            df = df.drop(columns=[c], errors="ignore")

    return df


def remove_total_or_blank_rows(df):
    if df.empty:
        return df

    df = df.copy().dropna(how="all").replace(["nan", "None", "NaT", "nat"], "")

    invoice_col = get_invoice_col(df)
    client_col = "Client" if "Client" in df.columns else None
    amount_col = "Amount" if "Amount" in df.columns else None

    if amount_col:
        df[amount_col] = df[amount_col].apply(clean_amount)

        required_cols = [
            c for c in ["Supplier Name", "Client", "Invoice number", "Currency"]
            if c in df.columns
        ]

        if required_cols:
            mask_total_only = (
                df[required_cols]
                .fillna("")
                .astype(str)
                .apply(lambda row: all(v.strip() == "" for v in row), axis=1)
            ) & df[amount_col].notna()

            df = df[~mask_total_only]

    if invoice_col and client_col:
        df = df[
            ~(
                df[invoice_col].fillna("").astype(str).str.strip().eq("")
                & df[client_col].fillna("").astype(str).str.strip().eq("")
            )
        ]

    return df.reset_index(drop=True)


def standardize_tracker_sheet(df, sheet_name):
    standard_cols = STANDARD_COLUMNS.get(sheet_name, [])

    if df.empty:
        return pd.DataFrame(columns=standard_cols)

    df = remove_total_or_blank_rows(clean_text_columns(remove_duplicate_columns(df.copy())))

    if "Supplier Name" in df.columns:
        df["Supplier Name"] = df["Supplier Name"].apply(normalize_supplier_name)

    if "Client" in df.columns:
        df["Client"] = df["Client"].apply(normalize_client)

    currency_col = next((c for c in df.columns if str(c).strip().lower() == "currency"), None)

    if currency_col:
        if currency_col != "Currency":
            df = df.rename(columns={currency_col: "Currency"})
        df["Currency"] = df["Currency"].astype(str).str.strip().str.upper()

    if "Amount" in df.columns:
        df["Amount"] = df["Amount"].apply(clean_amount)

    for col in ["Month of Service", "Month of Billing"]:
        if col in df.columns:
            df[col] = df[col].apply(normalize_month_value)

    if "Year" not in df.columns:
        df.insert(0, "Year", YEAR)
    else:
        df["Year"] = df["Year"].apply(lambda x: YEAR if pd.isna(x) or str(x).strip() == "" else x)

    for col in standard_cols:
        if col not in df.columns:
            df[col] = ""

    extra_cols = [c for c in df.columns if c not in standard_cols]
    return df[standard_cols + extra_cols].reset_index(drop=True)


def read_pdf_text(pdf_path):
    full_text = ""
    all_tables = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            full_text += page.extract_text() or ""
            all_tables += page.extract_tables() or []

    return full_text, all_tables


def detect_supplier_from_content(text, filename):
    t = text.upper()
    fname = os.path.basename(filename).upper()

    rules = [
        ("adsjoy", ["ADSJOY DIGITAL", "ADSJOY"]),
        ("apple", ["APPLE DISTRIBUTION", "APPLE SERVICES LATAM", "APPLE SEARCH ADS"]),
        ("meta", ["FACEBOOK", "META PLATFORMS"]),
        ("google", ["GOOGLE ADS", "PT GOOGLE", "GOOGLE LLC", "GOOGLE IRELAND", "GOOGLE ASIA PACIFIC", "COLLECTIONS@GOOGLE.COM"]),
    ]

    for supplier, keywords in rules:
        if any(k in t for k in keywords):
            return supplier

    if "ADSJOY" in fname:
        return "adsjoy"
    if re.match(r"Q\d+", fname):
        return "apple"
    if re.match(r"\d{10}(\s|\.|\-|$)", fname):
        return "google"
    if "TRANSACTION" in fname or re.match(r"\d{12,}", fname):
        return "meta"

    return "unknown"


def scan_invoices(root_folder):
    results = []

    for dirpath, _, files in os.walk(root_folder):
        for filename in sorted(files):
            if not filename.lower().endswith(".pdf"):
                continue

            pdf_path = os.path.join(dirpath, filename)

            try:
                text, tables = read_pdf_text(pdf_path)
            except Exception as e:
                log("ERR", f"Could not read {filename}: {e}", indent=2)
                text, tables = "", []

            results.append((pdf_path, detect_supplier_from_content(text, pdf_path), text, tables))

    return results


def enrich_with_gemini(base, text, supplier, pdf_path):
    if not gemini_available():
        return base

    try:
        enriched = enrich_extraction(base, text, supplier, pdf_path=pdf_path) or {}
        result = base.copy()

        for key, value in enriched.items():
            if value in [None, ""]:
                continue

            current = result.get(key)

            if current in [None, "", "UNKNOWN"] or str(current).strip().upper() in {"UNKNOWN", "BANK"}:
                result[key] = value

        return result

    except Exception as e:
        log("WARN", f"Gemini enrichment failed for {os.path.basename(pdf_path)}: {e}", indent=2)
        return base


def make_row(supplier, **kwargs):
    month = normalize_month_value(kwargs.get("month", ""))

    row = {
        "Year": YEAR,
        "Supplier Name": normalize_supplier_name(supplier),
        "Month of Service": month,
        "Month of Billing": normalize_month_value(kwargs.get("billing_month", month)),
        "Client": normalize_client(kwargs.get("client", "")),
        "Invoice number": str(kwargs.get("invoice_number", "") or "").strip(),
        "Currency": str(kwargs.get("currency", "") or "").strip().upper(),
        "Amount": clean_amount(kwargs.get("amount")),
    }

    for field in ["Market", "Ad Account ID", "Ad Account Name", "Campaign", "Campaign ID"]:
        if field in kwargs:
            row[field] = str(kwargs.get(field, "") or "").strip()

    return row


def filter_rows_against_existing(tracker_df, rows, supplier):
    if not rows:
        return []

    if tracker_df is None or tracker_df.empty:
        return rows

    filtered = []

    for row in rows:
        inv = str(row.get("Invoice number", "") or "").strip()
        campaign = str(row.get("Campaign", "") or "").strip()
        amount = clean_amount(row.get("Amount"))
        currency = str(row.get("Currency", "") or "").strip().upper()

        if not inv or not is_reasonable_amount(amount, currency):
            if inv:
                log("SKIP", f"{supplier}: invalid amount | invoice {inv} | {currency} {amount}", indent=2)
            continue

        exists = invoice_exists(tracker_df, inv)
        has_campaign_rows = invoice_has_campaign_rows(tracker_df, inv)

        if supplier == "adsjoy" and exists:
            log("SKIP", f"AdsJoy invoice {inv} already exists", indent=2)
            continue

        if supplier in {"apple", "google"} and exists:
            log("SKIP", f"{SUPPLIER_LABELS[supplier]} invoice {inv} already exists", indent=2)
            continue

        if supplier == "meta" and exists and not campaign:
            log("SKIP", f"Meta blank-campaign row skipped for existing invoice {inv}", indent=2)
            continue

        if exists and has_campaign_rows and not campaign:
            log("SKIP", f"Blank campaign total row skipped for existing invoice {inv}", indent=2)
            continue

        filtered.append(row)

    return filtered


def parse_adsjoy_pdf(pdf_path, text, tables, tracker_df=None):
    fname = os.path.basename(pdf_path)

    invoice_number = first_match([
        r"Invoice[:\s#]*([0-9]{2}-[0-9]{2}/[A-Za-z]{3}/[0-9]+)",
        r"Invoice\s*(?:No\.?|Number|#)?\s*[:\s]*([0-9\-/A-Za-z]+(?:/[0-9A-Za-z]+)*)",
    ], text)

    if invoice_exists(tracker_df, invoice_number):
        log("SKIP", f"AdsJoy invoice {invoice_number} already exists", indent=2)
        return []

    client = first_match([
        r"For\s+ADSJOY\s+DIGITAL\s*\n([A-Z]{2,6})\s",
        r"\$[\d,]+\.00\s*\nFor ADSJOY DIGITAL\s*\n([A-Z]{2,6})",
        r"\n([A-Z]{2,6})\s*\nFor\s+ADSJOY\s+DIGITAL",
    ], text)

    if not client:
        client = first_match(r"SAATCHI_([A-Z]+)_", fname).upper() or "UNKNOWN"

    if client.upper() in {"BANK", "ACCOUNT", "INVOICE", "TOTAL"}:
        client = "UNKNOWN"

    month = extract_month_from_text_or_filename(text, fname)
    currency = extract_currency(text, "USD")
    amount = None

    for table in tables:
        for row in table:
            row_text = " ".join(str(c) for c in row if c)
            if re.search(r"total|grand total|amount due", row_text, re.I):
                candidates = [
                    clean_amount(n)
                    for n in re.findall(r"[\d,]+\.?\d*", row_text)
                    if clean_amount(n) and clean_amount(n) > 100
                ]
                if candidates:
                    amount = max(candidates)
                    break
        if amount:
            break

    if amount is None:
        amount = clean_amount(first_match([
            r"TOTAL\s*\n[^\n]*\n\$?([\d,]+\.?\d*)",
            r"(?:Total|Grand Total|Amount Due)[^\n$]*\$?\s*([\d,]+\.?\d*)",
        ], text))

    enriched = enrich_with_gemini(
        {"client": client, "month": month, "currency": currency, "amount": amount},
        text,
        "adsjoy",
        pdf_path,
    )

    client = enriched.get("client", client)
    month = enriched.get("month", month)
    currency = enriched.get("currency", currency)
    amount = amount if amount is not None else clean_amount(enriched.get("amount"))

    if not is_reasonable_amount(amount, currency):
        log("SKIP", f"AdsJoy invoice {invoice_number} unreasonable amount: {currency} {amount}", indent=2)
        return []

    log("OK", f"AdsJoy invoice {invoice_number} | Client: {client} | {currency} {amount}", indent=2)

    return [make_row(
        "AdsJoy",
        month=month,
        client=client,
        invoice_number=invoice_number,
        campaign=client,
        campaign_id="",
        currency=currency,
        amount=amount,
    )]


def parse_apple_pdf(pdf_path, text, tables, tracker_df=None):
    fname = os.path.basename(pdf_path)

    invoice_number = first_match(r"(Q\d+)", fname, flags=re.I).upper()
    if not invoice_number:
        invoice_number = first_match(r"Invoice\s*Number\s*[:\s]*([A-Z0-9]+)", text) or "UNKNOWN"

    if invoice_exists(tracker_df, invoice_number):
        log("SKIP", f"Apple invoice {invoice_number} already exists", indent=2)
        return []

    client = first_match([
        r"Client\s*[:\s]+([A-Z]{2,6})\b",
        r"Order\s*Number\s*[:\s]*([A-Z]{2,6})\d*",
        r"Description\s*[:\s]*\(S\)[^\n]+\s+([A-Z]{2,6})\s*$",
    ], text, flags=re.I | re.MULTILINE).upper()

    if client in {"NAME", "AND", "ADDRESS", "NUMBER", "ID", "THE", ""}:
        client = "UNKNOWN"

    month = extract_month_from_text_or_filename(text, fname)
    currency = first_match(r"Currency\s*[:\s]*(USD|SGD|MYR|IDR|AUD|GBP|PHP)", text).upper() or extract_currency(text, "USD")
    market = first_match(r"\b(SG|MY|ID|TH|PH|MX|AU|GB|US)\b", text).upper()

    amount = clean_amount(first_match([
        r"Payable\s*Amount\s*\([A-Z]+\)\s*([\d,]+\.\d{2})",
        r"\bTotal\b\s+([\d,]+\.\d{2})",
        r"Subtotal\s+([\d,]+\.\d{2})",
    ], text))

    enriched = enrich_with_gemini(
        {"client": client, "market": market, "month": month, "currency": currency, "amount": amount},
        text,
        "apple",
        pdf_path,
    )

    client = enriched.get("client", client)
    market = enriched.get("market", market)
    month = enriched.get("month", month)
    currency = enriched.get("currency", currency)
    amount = amount if amount is not None else clean_amount(enriched.get("amount"))

    if not is_reasonable_amount(amount, currency):
        log("SKIP", f"Apple invoice {invoice_number} unreasonable amount: {currency} {amount}", indent=2)
        return []

    log("OK", f"Apple invoice {invoice_number} | Client: {client} | Market: {market} | {currency} {amount}", indent=2)

    return [make_row(
        "Apple (ASA)",
        month=month,
        client=client,
        market=market,
        invoice_number=invoice_number,
        currency=currency,
        amount=amount,
    )]


def parse_google_pdf(pdf_path, text, tables, tracker_df=None):
    fname = os.path.basename(pdf_path)

    invoice_number = first_match(r"Invoice\s*number[:\s.]*(\d+)", text) or first_match(r"(\d{10})", fname) or "UNKNOWN"

    if invoice_exists(tracker_df, invoice_number):
        log("SKIP", f"Google invoice {invoice_number} already exists", indent=2)
        return []

    month = extract_month_from_text_or_filename(text, fname)
    currency = extract_currency(text, "USD")

    client = first_match(r"(?:MCSP|mcsp)_[A-Z]{2}_([A-Z0-9]+)_", text).upper()
    if not client:
        client = first_match(r"^Account:\s*([A-Za-z0-9]+)", text, flags=re.I | re.MULTILINE).upper()
    if not client:
        client = "UNKNOWN"

    amount = clean_amount(first_match([
        r"Total\s+amount\s+due\s+in\s+[A-Z]{3}\s+[A-Z]{3}\s*([\d,]+\.\d{2})",
        r"Total\s+amount\s+due\s+in\s+[A-Z]{3}\s+[A-Z]{3}\s*([\d,]+)",
        r"Total\s+in\s+[A-Z]{3}\s+[A-Z]{3}\s*([\d,]+\.\d{2})",
        r"Total\s+in\s+[A-Z]{3}\s+[A-Z]{3}\s*([\d,]+)",
        r"Amount\s+due\s+[\w\s]*\s+([\d,]+\.\d{2})",
    ], text))

    enriched = enrich_with_gemini(
        {"client": client, "month": month, "currency": currency, "amount": amount},
        text,
        "google",
        pdf_path,
    )

    client = enriched.get("client", client)
    month = enriched.get("month", month)
    currency = enriched.get("currency", currency)

    if amount is None:
        enriched_amount = clean_amount(enriched.get("amount"))
        amount = enriched_amount if is_reasonable_amount(enriched_amount, currency) else None

    if amount is None or not is_reasonable_amount(amount, currency):
        log("SKIP", f"Google invoice {invoice_number}: no reliable total amount found", indent=2)
        return []

    log("OK", f"Google invoice {invoice_number} | Client: {client} | {currency} {amount} | {month}", indent=2)

    return [make_row(
        "Google",
        month=month,
        client=client,
        invoice_number=invoice_number,
        currency=currency,
        amount=amount,
    )]


def parse_meta_pdf(pdf_path, text, tables, tracker_df=None):
    invoice_number = first_match(r"Invoice\s*#[:\s]*(\d+)", text)
    period = first_match(r"Billing\s*Period[:\s]*([A-Za-z]{3}-\d{2})", text)
    billing_period = billing_period_to_str(period) if period else ""

    currency = first_match(r"Invoice\s*Currency[:\s]*(USD|SGD|MYR|IDR|AUD|GBP)", text).upper() or extract_currency(text, "")

    client = ""
    market = ""

    m = re.search(r"MCSP_([A-Z]{2})_([A-Z0-9]+)_", text, re.I)
    if m:
        market = m.group(1).upper()
        client = m.group(2).upper()

    if not client:
        m = re.search(r"mcspapac_([A-Z]{2})_[A-Z0-9]+_[^_]+_[A-Z]+_([A-Z]{2,6})_", text, re.I)
        if m:
            market = market or m.group(1).upper()
            client = m.group(2).upper()

    if not market:
        upper_text = text.upper()
        market = next((v for k, v in META_ENTITY_MARKET.items() if k in upper_text), "")

    if not client:
        account_id = first_match([
            r"Account\s*Id\s*/\s*Group[:\s]*(\d+)",
            r"(?<!\d)(\d{13,15})(?!\d)",
        ], text)
        client = META_ACCOUNT_CLIENT.get(account_id, "")

    if not client:
        client = first_match(r"Advertiser[:\s]*([^\n]+)", text) or "UNKNOWN"
        if re.search(r"saatchi|m&c|mcsaatchi", client, re.I):
            client = "UNKNOWN"

    invoice_total = clean_amount(first_match(r"Invoice\s*Total[:\s]*([\d,]+\.\d{2})", text))
    subtotal = clean_amount(first_match(r"Subtotal[:\s]*([\d,]+\.\d{2})", text))

    enriched = enrich_with_gemini(
        {"client": client, "market": market, "month": billing_period, "currency": currency, "amount": invoice_total},
        text,
        "meta",
        pdf_path,
    )

    client = enriched.get("client", client)
    market = enriched.get("market", market)
    billing_period = enriched.get("month", billing_period)
    currency = enriched.get("currency", currency)

    if invoice_total is None:
        invoice_total = clean_amount(enriched.get("amount"))

    rows = []

    for match in re.finditer(r"^\d+\s+(.+?)\s+([\d,]+\.\d{2})\s*$", text, re.MULTILINE):
        campaign_name = match.group(1).strip()
        amount = clean_amount(match.group(2))

        if (
            not campaign_name
            or amount is None
            or not is_reasonable_amount(amount, currency)
            or re.search(r"subtotal|invoice total|freight|vat|gst", campaign_name, re.I)
        ):
            continue

        rows.append(make_row(
            "Meta (Facebook)",
            month=billing_period,
            client=client,
            market=market,
            invoice_number=invoice_number,
            campaign=campaign_name,
            campaign_id=first_match(r"<([A-Z0-9]+)>", campaign_name),
            currency=currency,
            amount=amount,
            **{"Ad Account ID": ""},
        ))

    if not rows and not invoice_exists(tracker_df, invoice_number):
        fallback_amount = subtotal or invoice_total
        if is_reasonable_amount(fallback_amount, currency):
            rows.append(make_row(
                "Meta (Facebook)",
                month=billing_period,
                client=client,
                market=market,
                invoice_number=invoice_number,
                campaign="",
                campaign_id="",
                currency=currency,
                amount=fallback_amount,
                **{"Ad Account ID": ""},
            ))

    rows = filter_rows_against_existing(tracker_df, rows, "meta")

    log(
        "OK",
        f"Meta invoice {invoice_number} | Client: {client} | Market: {market} | {currency} | {len(rows)} row(s) | Total: {invoice_total}",
        indent=2,
    )

    return rows


def make_border(color="D9D9D9"):
    thin = Side(style="thin", color=color)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def format_sheet(ws, palette_key, amount_col_names=None):
    pal = COLOURS[palette_key]
    border = make_border()
    amount_col_names = amount_col_names or []

    header_fill = PatternFill("solid", fgColor=pal["header_bg"])
    alt_fill = PatternFill("solid", fgColor=pal["row_alt"])
    white_fill = PatternFill("solid", fgColor=WHITE)

    amount_cols = {
        ci for ci in range(1, ws.max_column + 1)
        if ws.cell(1, ci).value in amount_col_names
    }

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 30

    for ri in range(2, ws.max_row + 1):
        fill = alt_fill if ri % 2 == 0 else white_fill

        for ci in range(1, ws.max_column + 1):
            cell = ws.cell(ri, ci)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            if ci in amount_cols:
                cell.number_format = "#,##0.00"

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    for ci in range(1, ws.max_column + 1):
        col_letter = get_column_letter(ci)
        max_len = max(len(str(ws.cell(r, ci).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 55)


def build_summary(ws, sheet_map):
    border = make_border("CCCCCC")
    thin_border = make_border("E0E0E0")

    ws.merge_cells("A1:G1")
    ws["A1"] = "ACCT-108 | Master Invoice Tracker 2026 - Summary"
    ws["A1"].font = Font(bold=True, color=DARK, size=13)
    ws["A1"].fill = PatternFill("solid", fgColor="F0F4F8")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = border
    ws.row_dimensions[1].height = 36

    headers = ["Currency", "Supplier", "Sheet", "Clients", "# Invoices", "# Campaigns/Rows", "Total Amount"]

    for ci, header in enumerate(headers, 1):
        cell = ws.cell(2, ci, header)
        cell.fill = PatternFill("solid", fgColor="2E4057")
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    rows = []

    for sheet_name, df in sheet_map.items():
        if df.empty:
            continue

        df = standardize_tracker_sheet(df, sheet_name)

        if "Amount" not in df.columns or "Currency" not in df.columns:
            continue

        df["Amount"] = df["Amount"].apply(clean_amount)
        df = df[df["Amount"].notna() & df["Currency"].fillna("").astype(str).str.strip().ne("")].copy()

        if df.empty:
            continue

        for currency, grp in df.groupby("Currency"):
            currency = str(currency).strip().upper()
            invoice_col = get_invoice_col(grp)

            clients = "-"
            if "Client" in grp.columns:
                client_values = grp["Client"].dropna().astype(str).str.strip().apply(normalize_client)
                client_values = sorted([c for c in client_values.unique() if c and c.lower() not in {"nan", "none", "null", "nat"}])
                clients = ", ".join(client_values) if client_values else "-"

            invoice_count = (
                grp[invoice_col].astype(str).str.strip().replace("", pd.NA).dropna().nunique()
                if invoice_col else len(grp)
            )

            rows.append({
                "currency": currency,
                "supplier": {
                    "Adsjoy": "AdsJoy",
                    "Apple (ASA)": "Apple (ASA)",
                    "Google": "Google",
                    "Meta (facebook)": "Meta (Facebook)",
                }.get(sheet_name, sheet_name),
                "sheet": sheet_name,
                "clients": clients,
                "invoices": invoice_count,
                "rows": len(grp),
                "total": grp["Amount"].sum(),
            })

    order = ["IDR", "MYR", "SGD", "USD", "GBP"]
    rows.sort(key=lambda r: (order.index(r["currency"]) if r["currency"] in order else 99, r["supplier"]))

    ri = 3
    last_currency = None
    row_counter = 0

    for row in rows:
        currency = row["currency"]

        if currency != last_currency:
            color = CURRENCY_HEADER.get(currency, "333333")
            ws.merge_cells(f"A{ri}:G{ri}")
            cell = ws.cell(ri, 1, f"  {currency} - {currency} Invoices")
            cell.fill = PatternFill("solid", fgColor=color)
            cell.font = Font(bold=True, color=WHITE, size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cell.border = make_border(color)
            ri += 1
            row_counter = 0
            last_currency = currency

        row_counter += 1
        fill = PatternFill("solid", fgColor=CURRENCY_COLOURS.get(currency, "F5F5F5") if row_counter % 2 == 0 else WHITE)

        values = [currency, row["supplier"], row["sheet"], row["clients"], row["invoices"], row["rows"], row["total"]]

        for ci, value in enumerate(values, 1):
            cell = ws.cell(ri, ci, value)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.font = Font(color=DARK, size=10, bold=(ci == 7))
            if ci == 7:
                cell.number_format = "#,##0.00"

        ri += 1

    for ci, width in enumerate([10, 18, 18, 38, 13, 18, 18], 1):
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.freeze_panes = "A3"


def make_dedupe_key(df):
    if df.empty:
        return pd.Series(dtype=str)

    invoice_col = get_invoice_col(df)
    campaign_col = "Campaign" if "Campaign" in df.columns else None
    amount_col = "Amount" if "Amount" in df.columns else None
    currency_col = "Currency" if "Currency" in df.columns else None

    def safe(row, col):
        if not col or col not in row or pd.isna(row[col]):
            return ""
        return str(row[col]).strip()

    keys = []

    for _, row in df.iterrows():
        amt = clean_amount(row.get(amount_col)) if amount_col else None
        keys.append("|".join([
            safe(row, invoice_col),
            safe(row, campaign_col),
            f"{amt:.2f}" if amt is not None else "",
            safe(row, currency_col),
        ]))

    return pd.Series(keys, index=df.index)


def append_new_rows(existing_df, new_rows):
    stats = {"added": 0, "duplicates": 0}

    if not new_rows:
        return existing_df, stats

    new_df = pd.DataFrame(new_rows)

    if existing_df.empty:
        stats["added"] = len(new_df)
        return new_df.reset_index(drop=True), stats

    existing_keys = set(make_dedupe_key(existing_df).astype(str))
    new_keys = make_dedupe_key(new_df).astype(str)

    filtered = new_df[~new_keys.isin(existing_keys)].copy()

    stats["duplicates"] = len(new_df) - len(filtered)
    stats["added"] = len(filtered)

    return pd.concat([existing_df, filtered], ignore_index=True).reset_index(drop=True), stats


def load_tracker():
    log("LOAD", "Loading tracker from Input/...")

    tracker_sheets = pd.read_excel(TRACKER_PATH, sheet_name=None, header=0)
    tracker = {}

    for supplier, sheet_name in SHEET_NAMES.items():
        tracker[supplier] = standardize_tracker_sheet(
            tracker_sheets.get(sheet_name, pd.DataFrame()),
            sheet_name,
        )

    return tracker


def write_tracker(sheet_map):
    log("WRITE", "Writing updated tracker to Output/...")

    cleaned_sheet_map = {
        sheet_name: standardize_tracker_sheet(df, sheet_name)
        for sheet_name, df in sheet_map.items()
    }

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        pd.DataFrame().to_excel(writer, sheet_name="Summary", index=False)

        for sheet_name, df in cleaned_sheet_map.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    log("FORMAT", "Applying workbook formatting...")

    wb = load_workbook(OUTPUT_PATH)

    for sheet_name, config in {
        "Adsjoy": ("adsjoy", ["Amount"]),
        "Apple (ASA)": ("apple", ["Amount"]),
        "Google": ("google", ["Amount"]),
        "Meta (facebook)": ("meta", ["Amount"]),
    }.items():
        if sheet_name in wb.sheetnames:
            format_sheet(wb[sheet_name], config[0], amount_col_names=config[1])

    build_summary(wb["Summary"], cleaned_sheet_map)

    if "Summary" in wb.sheetnames:
        wb.move_sheet("Summary", offset=-(len(wb.sheetnames) - 1))

    wb.save(OUTPUT_PATH)


def print_extraction_summary(stats, output_path, elapsed):
    print()
    print(SUB_SEP)
    print("EXTRACTION SUMMARY")
    print(SUB_SEP)
    print(f"PDFs scanned          : {stats['pdfs_scanned']}")
    print(f"Unknown skipped       : {stats['unknown_skipped']}")
    print(f"Rows extracted        : {stats['rows_extracted']}")
    print(f"Rows added            : {stats['rows_added']}")
    print(f"Duplicates skipped    : {stats['duplicates']}")
    print()

    for supplier in ["adsjoy", "apple", "google", "meta"]:
        s = stats["suppliers"][supplier]
        print(f"{SUPPLIER_LABELS[supplier]:<20}: PDFs {s['pdfs']:<3} | Rows {s['rows']:<4} | Added {s['added']:<4} | Duplicates {s['duplicates']}")

    print()
    print(f"Output file           : {output_path}")
    print(f"Duration              : {duration_str(elapsed)}")
    print("Result                : SUCCESS")
    print(SUB_SEP)


def main():
    start = perf_counter()

    print_header()

    os.makedirs(INPUT_FOLDER, exist_ok=True)
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    if not os.path.exists(TRACKER_PATH):
        print()
        log("ERR", f"Tracker not found: {TRACKER_PATH}")
        log("INFO", "Please place your tracker in the Input/ folder and re-run.")
        sys.exit(1)

    tracker = load_tracker()
    new_rows = {supplier: [] for supplier in SHEET_NAMES}

    stats = {
        "pdfs_scanned": 0,
        "unknown_skipped": 0,
        "rows_extracted": 0,
        "rows_added": 0,
        "duplicates": 0,
        "suppliers": {
            supplier: {"pdfs": 0, "rows": 0, "added": 0, "duplicates": 0}
            for supplier in SHEET_NAMES
        },
    }

    parsers = {
        "adsjoy": parse_adsjoy_pdf,
        "apple": parse_apple_pdf,
        "google": parse_google_pdf,
        "meta": parse_meta_pdf,
    }

    invoices = scan_invoices(INVOICE_FOLDER)
    stats["pdfs_scanned"] = len(invoices)

    print()
    log("SCAN", f"Found {len(invoices)} PDF invoice(s) across all subfolders")
    print()

    for pdf_path, supplier, text, tables in invoices:
        filename = os.path.basename(pdf_path)
        print(f"[PDF] {filename} -> [{supplier.upper()}]")

        if supplier == "unknown":
            stats["unknown_skipped"] += 1
            log("WARN", "Unknown supplier — skipped", indent=2)
            continue

        stats["suppliers"][supplier]["pdfs"] += 1

        try:
            result = parsers[supplier](pdf_path, text, tables, tracker[supplier])
            result = filter_rows_against_existing(tracker[supplier], result, supplier)

            if result:
                new_rows[supplier].extend(result)
                stats["rows_extracted"] += len(result)
                stats["suppliers"][supplier]["rows"] += len(result)

        except Exception as e:
            log("ERR", f"Failed to parse {filename}: {e}", indent=2)

    print()
    log("MERGE", "Merging extracted rows into tracker sheets...")

    sheet_map = {}

    for supplier, sheet_name in SHEET_NAMES.items():
        merged_df, merge_stats = append_new_rows(tracker[supplier], new_rows[supplier])
        sheet_map[sheet_name] = standardize_tracker_sheet(merged_df, sheet_name)

        stats["rows_added"] += merge_stats["added"]
        stats["duplicates"] += merge_stats["duplicates"]
        stats["suppliers"][supplier]["added"] = merge_stats["added"]
        stats["suppliers"][supplier]["duplicates"] = merge_stats["duplicates"]

        log("MERGE", f"{SUPPLIER_LABELS[supplier]}: {merge_stats['added']} added, {merge_stats['duplicates']} duplicate(s) skipped", indent=2)

    write_tracker(sheet_map)

    elapsed = perf_counter() - start

    print()
    log("DONE", "Extraction complete")
    print_extraction_summary(stats, OUTPUT_PATH, elapsed)
    print(SEP)


if __name__ == "__main__":
    main()